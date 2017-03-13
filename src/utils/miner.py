#!/usr/bin/python
# coding=utf-8

from meta import *
from fancyimpute import *
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy.matlib
import numpy
import time


# A Miner is responsible for mining the correlation between measure data human
# input:  all original data containing: measure...
# output: all correlation between measures
# usage:  impute the missing value in measure data
class Miner:
    __metaclass__ = Singleton

    def __init__(self, data):
        # load all necessary data
        self.test_size = 300
        self.data = data
        self.paras = self.data.paras
        self.flag_ = self.data.flag_

        self.GIRTH = numpy.array(self.data.paras["GIRTH"], dtype=bool)
        self.GIRTH.shape = (self.data.m_num, 1)
        self.LENGTH = numpy.array(self.data.paras["LENGTH"], dtype=bool)
        self.LENGTH.shape(self.data.m_num, 1)

        self.data_path = self.paras['dataPath'] + "miner/"
        self.m2m = list(self.paras['m2m'])
        [self.m_basis, self.m_coeff, self.m_pca_mean, self.m_pca_std] = \
            self.get_m_basis()
        self.cfTable = self.itemCF()

        self.impute_name = ["SimpleFill", "SimpleAverage",
                            "CaseAmplifify" "KNN", "MICE"]
        self.impute_method = [SimpleFill(), self.simpleAverage,
                              self.caseAmplify, KNN, MICE()]
        print(type(self.simpleAverage))

    # given flag, value, predict completed measures
    def getPredict(self, flag, data):
        if (flag == 1).sum() == self.data.m_num:
            return data
        else:
            solver = MICE()
            return self.Imputate(flag, data, True, solver)

    # calculating measure-based presentation(PCA)
    def get_m_basis(self):
        print(" [**] begin get_measure_basis ...")
        m_basis_file = self.data_path + 'm_basis_0%d.npy' % self.flag_
        m_coeff_file = self.data_path + 'm_coeff_0%d.npy' % self.flag_
        m_pca_mean_file = self.data_path + 'm_pca_mean_0%d.npy' % self.flag_
        m_pca_std_file = self.data_path + 'm_pca_std_0%d.npy' % self.flag_
        start = time.time()
        if self.data.paras["reload_m_basis"]:
            # principle component analysis
            t_measure = self.data.measure - self.data.mean_measure
            t_measure /= self.data.std_measure
            m_basis, sigma, M = numpy.linalg.svd(t_measures, full_matrices=0)
            m_coeff = numpy.dot(m_basis.transpose(), t_measure)
            m_pca_mean = numpy.array(numpy.mean(m_coeff, axis=1))
            m_pca_mean.shape = (m_pca_mean.size, 1)
            m_pca_std = numpy.array(numpy.std(m_coeff, axis=1))
            m_pca_std.shape = (m_pca_std.size, 1)
            m_basis = numpy.array(m_basis)
            m_coeff = numpy.array(m_coeff).reshape(
                m_coeff.shape[0], self.data.body_count)
            numpy.save(open(m_basis_file, "wb"), m_basis)
            numpy.save(open(m_coeff_file, "wb"), m_coeff)
            numpy.save(open(m_pca_mean_file, "wb"), m_pca_mean)
            numpy.save(open(m_pca_std_file, "wb"), m_pca_std)
        else:
            m_basis = numpy.load(open(m_basis_file, "rb"))
            m_coeff = numpy.load(open(m_coeff_file, "rb"))
            m_pca_mean = numpy.load(open(m_pca_mean_file, "rb"))
            m_pca_std = numpy.load(open(m_pca_std_file, "rb"))
        print(' [**] finish get_m_basis in %fs' % (time.time() - start))
        return [m_basis, m_sigma, m_pca_mean, m_pca_std, m_coeff]

    # cal the similarity between measure, using Pearson Correlation(300-1531)
    def itemCF(self):
        print(' [**] begin get similarity of measures')
        start = time.time()
        cf_path = self.data_path + "cfTable.npy"
        table = numpy.zeros((self.data.m_num, self.data.m_num))
        if self.data.paras['reload_cfTable']:
            # calculation
            t_measure = self.data.measure - self.data.mean_measure
            t_measure /= self.data.std_measure
            for i in range(0, self.data.m_num):
                table[i, i] = 0
                for j in range(i + 1, self.data.m_num):
                    a = numpy.array(t_measure[i, self.test_size:])
                    a.shape = (1, self.data.body_num - self.test_size)
                    b = numpy.array(t_measure[j, self.test_size:])
                    b.shape(self.data.body_num - self.test_size, 1)
                    similar = numpy.linalg.norm(a) * numpy.linalg.norm(b)
                    similar = (a.dot(b) / similar)[0, 0]
                    table[i, j] = table[j, i] = similar
            # save the result in excel
            wb = Workbook()
            ws = wb.get_active_sheet()
            top = self.data.measure_num
            for i in range(0, top):
                ws.cell(row=i * (top + 2) + 1, column=1).value = i
                ws.cell(row=i * (top + 2) + 1,
                        column=2).value = self.data.measure_str[i]
                for j in range(0, top):
                    ws.cell(row=i * (top + 2) + j + 2,
                            column=1).value = self.data.measure_str[j]
                    ws.cell(row=i * (top + 2) + j + 2,
                            column=2).value = table[i, j]
            wb.save(filename=self.data.paras['ansPath'] + 'similarity.xlsx')
            numpy.save(open(cf_path, "wb"), table)
        else:
            table = numpy.load(open(cf_path, "rb"))
        print(' [**] finish load cfTable in %fs.' % (time.time() - start))
        return table

    # given flag value, predict completed measures, simple sum all
    def simpleAverage(self, flag, data, dist):
        data = numpy.array(data).reshape(self.data.measure_num, 1)
        output = flag * data

        for j in range(0, output.shape[0]):
            if output[j, 0] == 0:
                coeff = numpy.array(self.cfTable[j, :]).reshape(
                    self.data.measure_num, 1)
                if (abs(coeff).sum() == 0):
                    output[j, 0] = 0
                else:
                    if dist & ((self.GIRTH[j, 0] | self.LENGTH[j, 0]) != 0):
                        if self.GIRTH[j, 0]:
                            x = data[flag & self.GIRTH]
                            coeff = coeff[flag & self.GIRTH]
                        else:
                            x = data[flag & self.LENGTH]
                            coeff = coeff[flag & self.LENGTH]
                        x.shape = (1, x.size)
                    else:
                        x = data[flag]
                        x.shape = (1, x.size)
                        coeff = coeff[flag]
                    output[j, 0] = x.dot(coeff) / (abs(coeff).sum())
        return output

    # Predict, let small smaller, big bigger
    def caseAmplify(self, flag, data, dist):
        frac = 2.5
        data = numpy.array(data).reshape(self.data.measure_num, 1)
        output = flag * data

        for j in range(0, output.shape[0]):
            if output[j, 0] == 0:
                coeff = numpy.array(self.cfTable[j, :]).reshape(
                    self.data.measure_num, 1)
                if (abs(coeff).sum() == 0):
                    output[j, 0] = 0
                else:
                    if dist & ((self.GIRTH[j, 0] | self.LENGTH[j, 0]) != 0):
                        if self.GIRTH[j, 0]:
                            x = data[flag & self.GIRTH]
                            coeff = coeff[flag & self.GIRTH]
                        else:
                            x = data[flag & self.LENGTH]
                            coeff = coeff[flag & self.LENGTH]
                        x.shape = (1, x.size)
                    else:
                        x = data[flag]
                        x.shape = (1, x.size)
                        coeff = coeff[flag]
                    tmp = coeff * (abs(coeff)**1.5)
                    output[j, 0] = x.dot(tmp) / (abs(tmp).sum())
        return output

    # using imputation for missing data
    def Imputate(self, flag, data, dist, imputor):
        output = numpy.array(data).reshape(self.data.measure_num, 1)
        output[~flag] = numpy.nan
        tmp = numpy.array(self.data.t_measures[:, self.test_size:])
        tmp = numpy.column_stack((tmp, output)).transpose()
        tmp = imputor.complete(tmp)
        output = numpy.array(tmp[-1, :]).reshape(self.data.measure_num, 1)
        return output

    # test all methods
    def test(self):
        print(' [**] Begin predict input data...')
        table_size = 26
        input_file = self.filePath + "input.xlsx"
        input_wb = load_workbook(filename=input_file)
        sheets = input_wb.get_sheet_names()
        input_ws = input_wb.get_sheet_by_name(sheets[0])
        output_file = '../../result/prediction.xlsx'
        output_wb = Workbook()
        output_ws = output_wb.get_active_sheet()

        # read test data
        test_data = []
        for i in range(0, 5):
            flag = numpy.zeros((self.data.measure_num, 1), dtype=bool)
            for j in range(0, self.data.measure_num):
                flag[j, 0] = input_ws.cell(row=2 + j, column=2 + i).value
            test_data.append(flag)

        # begin test, using 0-300 data, output the offset to mean
        for i in range(0, len(test_data)):
            output_ws.cell(row=(i * table_size) + 2, column=1).value = i
            for j in range(0, self.data.measure_num):
                output_ws.cell(row=(i * table_size) + 3 + j,
                               column=1).value = self.data.measure_str[j]
            output_ws.cell(row=(i * table_size) + 22,
                           column=1).value = 'MEAN GIRTH'
            output_ws.cell(row=(i * table_size) + 23,
                           column=1).value = 'MEAN LENGTH'
            output_ws.cell(row=(i * table_size) + 24, column=1).value = 'MEAN'
            k = 2

            # test for imputation technology
            for m in range(0, len(self.impute_method)):
                output_ws.cell(row=(i * table_size) + 2,
                               column=k).value = self.impute_name[m]
                error = numpy.zeros((self.data.measure_num, 1))
                for j in range(0, self.test_size):
                    print('Experiment %d, ' %
                          i, self.impute_name[m], ' sample: ', j)
                    input = numpy.array(self.data.t_measures[:, j]).reshape(
                        self.data.measure_num, 1)
                    output = self.Imputate(
                        test_data[i], input, False, self.impute_method[m])
                    error += abs(output - input)
                error /= self.test_size
                error *= self.data.std_measures
                for j in range(0, self.data.measure_num):
                    output_ws.cell(row=(i * table_size) + 3 + j,
                                   column=k).value = error[j, 0]
                output_ws.cell(row=(i * table_size) + 22, column=k).value = numpy.average(
                    abs(error[(~test_data[i]) & self.GIRTH]))
                output_ws.cell(row=(i * table_size) + 23, column=k).value = numpy.average(
                    abs(error[(~test_data[i]) & self.LENGTH]))
                output_ws.cell(row=(
                    i * table_size) + 24, column=k).value = numpy.average(abs(error[(~test_data[i])]))
                k = k + 1

            # test for simple weighted without classification
            output_ws.cell(row=(i * table_size) + 2,
                           column=k).value = 'simple weighted'
            error = numpy.zeros((self.data.measure_num, 1))
            for j in range(0, self.test_size):
                input = numpy.array(self.data.t_measures[:, j]).reshape(
                    self.data.measure_num, 1)
                output = self.simpleAverage(test_data[i], input, False)
                error += abs(output - input)
            error /= self.test_size
            error *= self.data.std_measures
            for j in range(0, self.data.measure_num):
                output_ws.cell(row=(i * table_size) + 3 + j,
                               column=k).value = error[j, 0]
            output_ws.cell(row=(i * table_size) + 22, column=k).value = numpy.average(
                abs(error[(~test_data[i]) & self.GIRTH]))
            output_ws.cell(row=(i * table_size) + 23, column=k).value = numpy.average(
                abs(error[(~test_data[i]) & self.LENGTH]))
            output_ws.cell(row=(i * table_size) + 24,
                           column=k).value = numpy.average(abs(error[(~test_data[i])]))
            k += 1

            # test for case amplication(2.5) without classification
            output_ws.cell(row=(i * table_size) + 2,
                           column=k).value = 'case amplication'
            error = numpy.zeros((self.data.measure_num, 1))
            for j in range(0, self.test_size):
                input = numpy.array(self.data.t_measures[:, j]).reshape(
                    self.data.measure_num, 1)
                output = self.caseAmplify(test_data[i], input, False)
                error += abs(output - input)
            error /= self.test_size
            error *= self.data.std_measures
            for j in range(0, self.data.measure_num):
                output_ws.cell(row=(i * table_size) + 3 + j,
                               column=k).value = error[j, 0]
            output_ws.cell(row=(i * table_size) + 22, column=k).value = numpy.average(
                abs(error[(~test_data[i]) & self.GIRTH]))
            output_ws.cell(row=(i * table_size) + 23, column=k).value = numpy.average(
                abs(error[(~test_data[i]) & self.LENGTH]))
            output_ws.cell(row=(i * table_size) + 24,
                           column=k).value = numpy.average(abs(error[(~test_data[i])]))
            k = k + 1

            output_wb.save(output_file)
        output_wb.save(output_file)
        print(' [**] Finish predict input data...')


# test for this module
if __name__ == "__main__":
    male = rawData("../parameter.json", 1)
    female = rawData("../parameter.json", 2)

    male_miner = Miner(male)
    # male_miner.test()

    # female_miner = Miner(female)
    # female_miner.test()
