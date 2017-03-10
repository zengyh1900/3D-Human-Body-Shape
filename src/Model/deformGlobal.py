#!/usr/bin/python
#coding=utf-8

import sys
sys.path.append("..")
from dataProcess.rawData import *
from dataProcess.dataModel import *
from dataProcess.basisData import *
from openpyxl import Workbook
import scipy.sparse.linalg
import scipy.sparse
import scipy
import numpy as np
import time


class deformGlobal:
    '''
        a class used to calculate corrlated massage about deform-based global method
        given: deform-based PCA coefficient, measure data
        output: the matrix trans measure into deform-based PCA space
    '''
    __metaclass__ = Singleton
    def __init__(self, data):
        self.TYPE = "deform-global"
        self.data = data
        self.deformation = None
        self.D = self.load_global_matrix()
        self.demo_num = self.data.measure_num

    # ---------------------------------------------------------------------
    '''calculate global mapping from measure to deformation PCA coeff'''
    # ---------------------------------------------------------------------
    def load_global_matrix(self):
        print (' [**] begin d_global_mapping ... ')
        start = time.time()
        if self.data.paras['reload_D']:
            W = self.data.d_coeff[:self.data.d_basis_num, :]  # (10, 1531)
            W = W.transpose().reshape((W.size, 1))
            M = self.data.build_equation(self.data.t_measures, self.data.d_basis_num)
            # solve transform matrix
            MtM = M.transpose().dot(M)
            MtW = M.transpose().dot(W)
            D = np.array(scipy.sparse.linalg.spsolve(MtM, MtW)).reshape((self.data.d_basis_num, self.data.measure_num))
            np.save(open(self.data.NPYpath+'D.npy', 'wb'), D)
            print (' [**] finish d_global_mapping in %fs' % (time.time() - start))
            return D
        else:
            print (' [**] finish d_global_mapping in %fs' % (time.time() - start))
            return np.load(open(self.data.NPYpath+'D.npy', 'rb'))

    # ----------------------------------------------------------------
    '''rebuild the female dataset using deform-based global method'''
    # ----------------------------------------------------------------
    def global_rebuild(self):
        wb = Workbook()
        ws = wb.get_active_sheet()
        all = np.zeros((self.data.measure_num, self.data.body_count))
        for i in range(0, self.data.measure_num):
            ws.cell(row=1, column=i+2).value = self.data.measure_str[i]
        for i in range(0, self.data.body_count):
            print('rebuilding deform_global-based: %d  ...'%i)
            ws.cell(row=i+2, column=1).value = i
            input = self.data.t_measures[:, i].reshape(self.data.measure_num, 1)
            [vertex, n, f] = self.mapping(input)
            # self.data.save_obj(self.data.ansPath+self.data.o_file_list[i], vertex, f+1)

            input = self.data.mean_measures + self.data.std_measures * input
            output = np.array(self.data.calc_measures(vertex))
            error = output - input
            error[0,0] = (output[0,0]**3)/(1000**3) - (input[0,0]**3)/(1000**3)
            all[:,i] = error.flat
            for j in range(0, error.shape[0]):
                ws.cell(row=i+2, column=j+2).value=error[j,0]
        std = np.std(all, axis=1)
        mean = np.mean(abs(all), axis = 1)
        ws.cell(row=self.data.body_count+2, column=1).value = "mean error"
        ws.cell(row=self.data.body_count+3, column=1).value = "std"
        for i in range(0, len(mean)):
            ws.cell(row=self.data.body_count+2, column=i+2).value = mean[i]
            ws.cell(row=self.data.body_count+3, column=i+2).value = std[i]
        wb.save(self.data.ansPath+'rebuild_d_global.xlsx')

    # -----------------------------------------------------------------------
    '''given t_measure, output vertex, using deform-based global method'''
    # -----------------------------------------------------------------------
    def mapping(self, data):
        data = np.array(data[:self.demo_num,:]).reshape(self.demo_num, 1)
        alpha = self.D.dot(data)
        basis = self.data.d_basis[:, :self.data.d_basis_num]
        alpha = np.array(alpha).reshape(alpha.size, 1)
        self.deformation = self.data.mean_deform + np.matmul(basis, alpha) * self.data.std_deform
        return self.data.d_synthesize(self.deformation)


#############################################
'''test'''
#############################################
if __name__ == "__main__":
    filename = "../parameter.json"
    data = rawData(filename)
    bd = basisData(data)
    mark = Masker(data)
    model = dataModel(bd, mark)

    dg = deformGlobal(model)
    dg.global_rebuild()

