#!/usr/bin/python
#coding=utf-8

import sys
sys.path.append("..")
from dataProcess.rawData import *
from dataProcess.dataModel import *
from openpyxl import Workbook
import numpy as np
import scipy.sparse.linalg
import scipy.sparse
import scipy
import time

class vertexGlobal:
    '''
        a vertex-based global model, mainly for synthesis body from measure
        input: body measures
        usage: build a model in vertex-based pca space, mapping measures to this space
        output: 3D human body shape
    '''
    def __init__(self, data):
        self.type = "vertex-global"
        self.data = data
        self.deformation = None
        self.V = self.load_global_matrix()
        self.demo_num = self.data.measure_num

    # ----------------------------------------------------------------
    '''calculate the mapping matrix from measures to vertex-based '''
    # ----------------------------------------------------------------
    def load_global_matrix(self):
        print (' [**] begin load_global_matrix ... ')
        start = time.time()
        if self.data.paras["reload_v_mapping"]:
            W = np.array(self.data.v_coeff[:self.data.v_basis_num,::]).reshape(self.data.v_basis_num, self.data.body_count)
            W = W.transpose().reshape((W.size, 1))
            M = self.data.build_equation(self.data.t_measures, self.data.v_basis_num)
            # solve transform matrix
            MtM = M.transpose().dot(M)
            MtW = M.transpose().dot(W)
            V = np.array(scipy.sparse.linalg.spsolve(MtM, MtW))
            V.shape = ((self.data.v_basis_num, self.data.measure_num))
            np.save(open(self.data.NPYpath+'V.npy', 'wb'), V)
            print (' [**] finish load_global_matrix between measures & vertex-based in %fs' % (time.time() - start))
            return V
        else:
            print (' [**] finish load_global_matrix between measures & vertex-based in %fs' % (time.time() - start))
            return np.load(open(self.data.NPYpath+'V.npy', 'rb'))

    # ----------------------------------------------------------------
    '''rebuild the female dataset by vertex-global method '''
    # ----------------------------------------------------------------
    def v_rebuild(self):
        wb = Workbook()
        ws = wb.get_active_sheet()
        all = np.zeros((self.data.measure_num, self.data.body_count))
        for i in range(0, self.data.measure_num):
            ws.cell(row=1, column=i+2).value = self.data.measure_str[i]
        for i in range(0, self.data.body_count):
            print('rebuilding vertex_global-based: %d  ...'%i)
            ws.cell(row=i + 2, column=1).value = i
            input = self.data.t_measures[:,i].reshape(self.data.measure_num, 1)
            [vertex, n, f] = self.mapping(input)
            # self.data.save_obj(self.data.ansPath+self.data.o_file_list[i],vertex,f+1)
            input = self.data.mean_measures + self.data.std_measures*input
            output = np.array(self.data.calc_measures(vertex))
            error = output-input
            error[0,0] = (output[0,0]**3)/(1000**3) - (input[0,0]**3)/(1000**3)
            all[:,i] = error.flat
            for j in range(0, error.shape[0]):
                ws.cell(row=i+2, column=j+2).value = error[j,0]
        std = np.std(all, axis=1)
        mean = np.mean(abs(all), axis = 1)
        ws.cell(row=self.data.body_count+2, column=1).value = "mean error"
        ws.cell(row=self.data.body_count+3, column=1).value = "std"
        for i in range(0, len(mean)):
            ws.cell(row=self.data.body_count+2, column=i+2).value = mean[i]
            ws.cell(row=self.data.body_count+3, column=i+2).value = std[i]
        wb.save(self.data.ansPath+'rebuild_v_global.xlsx')

    # -----------------------------------------------------------------------------------
    '''given t_measure, return body shape'''
    # -----------------------------------------------------------------------------------
    def mapping(self, measure):
        measure = np.array(measure[:self.demo_num, :]).reshape(self.demo_num, 1)
        weight = self.V.dot(measure)
        [v, n, f]  = self.data.v_synthesize(weight)
        # self.deformation = self.data.getDeform(v)
        return [v,n,f]

#############################################
'''test'''
#############################################
if __name__ == "__main__":
    filename = "../parameter.json"
    data = rawData(filename)
    bd = basisData(data)
    mark = Masker(data)
    model = dataModel(bd, mark)

    vg = vertexGlobal(model)
    vg.v_rebuild()