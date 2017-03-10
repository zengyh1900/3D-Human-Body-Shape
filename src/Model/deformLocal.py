#!/usr/bin/python
#coding=utf-8

import sys
sys.path.append("..")
from dataProcess.dataModel import *
from openpyxl import Workbook
import scipy.sparse.linalg
import numpy as np


class deformLocal:
    '''
        a model map measures to local face
    '''
    __metaclass__ = Singleton
    def __init__(self, data):
        self.TYPE = "deform-local"
        self.data = data
        self.deformation = None
        self.L_list = self.local_matrix()
        self.demo_num = self.data.measure_num

    # --------------------------------------------------------------------------
    '''local map matrix: measure->deform'''
    # --------------------------------------------------------------------------
    def local_matrix(self):
        print (' [**] begin solve local transformation')
        start = time.time()
        L_file = self.data.NPYpath + ("localMapper/L%d.npy" % (self.data.paras['mapping_version']))
        if self.data.paras['reload_L']:
            L_list = []
            L_tosave = []
            for i in range(0, self.data.face_num):
                print ('  calculating local mapping with face: %d' % i)
                S = np.array(self.data.o_deform[i * 9:i * 9 + 9, :]).reshape(9, self.data.body_count)
                S = S.transpose()
                S = S.reshape(S.size, 1)
                mask = np.array(self.data.mask[i * 19:i * 19 + 19, 0]).reshape(19, 1)
                mask = mask.repeat(1531, axis=1)
                measures = np.array(self.data.o_measures[mask])
                measures.shape = (measures.size / self.data.body_count, self.data.body_count)
                M = self.data.build_equation(measures, 9)
                # solve transform matrix
                MtM = M.transpose().dot(M)
                MtS = M.transpose().dot(S)
                L = np.array(scipy.sparse.linalg.spsolve(MtM, MtS))
                L.shape = (9, measures.size / self.data.body_count)
                # #==================================
                # s = np.array(self.data.o_deform[i * 9:i * 9 + 9, 0]).reshape(9, 1)
                # m = np.array(self.data.o_measures[:,0][mask[:,0]])
                # m.shape = (m.size, 1)
                # print('L: ', L.shape, 's:', s.shape, 'm:', m.shape)
                # print('L: \n', L)
                # print('before: \n', s)
                # print('back: \n', L.dot(m).reshape(9, 1))
                # input()
                # #=================================
                L_list.append(L)
                L_tosave.append(list(L))
            self.data.save_NPY([L_file],[L_tosave])
            print (' [**] finish solve local transformation in %fs' % (time.time() - start))
            return L_list
        else:
            print (' [**] finish solve local transformation in %fs' % (time.time() - start))
            tmp = self.data.load_NPY([L_file])[0]
            L_list = []
            for i in range(0, len(tmp)):
                L_list.append(np.array([c for c in tmp[i]]))
            return L_list

    # ----------------------------------------------------------------
    '''rebuild the female dataset using deform-based local method '''
    # ----------------------------------------------------------------
    def local_rebuild(self):
        wb = Workbook()
        ws = wb.get_active_sheet()
        all = np.zeros((self.data.measure_num, self.data.body_count))
        for i in range(0, self.data.measure_num):
            ws.cell(row=1, column=i+2).value = self.data.measure_str[i]
        for i in range(0, self.data.body_count):
            print('rebuilding deform_local-based_v%d: %d  ...'%(self.data.paras['mapping_version'], i))
            ws.cell(row=i+2, column=1).value = i
            input = self.data.t_measures[:, i].reshape(self.data.measure_num, 1)
            [vertex, n, f] = self.mapping(input)
            # self.data.save_obj(self.data.ansPath+self.data.o_file_list[i], vertex, f+1)

            input = self.data.mean_measures + self.data.std_measures * input
            output = np.array(self.data.calc_measures(vertex))
            error = output - input
            error[0,0] = (output[0,0]**3)/(1000**3) - (input[0,0]**3)/(1000**3)
            all[:,i] = error.flat
            for j in range(0, error.shape[0]):
                ws.cell(row=i+2, column=j+2).value=error[j,0]
        std = np.std(all, axis=1)
        mean = np.mean(abs(all), axis = 1)
        ws.cell(row=self.data.body_count+2, column=1).value = "mean error"
        ws.cell(row=self.data.body_count+3, column=1).value = "std"
        for i in range(0, len(mean)):
            ws.cell(row=self.data.body_count+2, column=i+2).value = mean[i]
            ws.cell(row=self.data.body_count+3, column=i+2).value = std[i]
        wb.save(self.data.ansPath+'rebuild_d_local_v%d.xlsx'%(self.data.paras['mapping_version']))

    # ------------------------------------------------------------------------------------
    '''given t_measures, return body shape '''
    # ------------------------------------------------------------------------------------
    def mapping(self, data):
        data = np.array(data[:self.demo_num,:]).reshape(self.demo_num, 1)
        data = self.data.mean_measures + self.data.std_measures * data
        deform = []
        for i in range(0, self.data.face_num):
            mask = np.array(self.data.mask[i * 19:i * 19 + 19, 0]).reshape(19, 1)
            alpha = np.array(data[mask])
            alpha.shape = (alpha.size, 1)
            s = self.L_list[i].dot(alpha)
            deform += [a for a in s.flat]
        deform = np.array(deform).reshape(len(deform), 1)
        self.deformation = deform
        return self.data.d_synthesize(deform)


#############################################
'''test'''
#############################################
if __name__ == "__main__":
    filename = "../parameter.json"
    data = rawData(filename)
    bd = basisData(data)
    mark = Masker(data)
    model = dataModel(bd, mark)

    dl = deformLocal(model)
    dl.local_rebuild()
