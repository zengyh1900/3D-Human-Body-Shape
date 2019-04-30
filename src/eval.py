
import numpy as np
import os
import time
import utils
from reshaper import Reshaper
from openpyxl import Workbook
from openpyxl import load_workbook
from multiprocessing import Pool, Process

ANS_DIR = "../Rebuild"
DATA_DIR = "../data"
label = "male"
flag = 2
body = Reshaper(label=label)

t_measure = np.load(open(os.path.join(DATA_DIR, "%s_t_measure.npy"%label), "rb"))
mean_measure = np.load(open(os.path.join(DATA_DIR, "%s_mean_measure.npy"%label), "rb"))
std_measure = np.load(open(os.path.join(DATA_DIR, "%s_std_measure.npy"%label), "rb"))
facet = np.load(open(os.path.join(DATA_DIR, "facet.npy"), "rb"))
cp = np.load(open(os.path.join(DATA_DIR, "cp.npy"), "rb"))
body_num = t_measure.shape[1]

def rebuild(idx):
  # print("[**] processing %d"%idx)
  data = t_measure[:, idx].reshape(utils.M_NUM, 1)
  [vertex, n, f] = body.mapping(data, flag=flag)
  utils.save_obj(os.path.join(ANS_DIR, "%s_%d.obj"%(label, idx)), vertex, f + 1)
  data = mean_measure + data * std_measure
  output = np.array(utils.calc_measure(cp, vertex, facet))
  error = output - data
  error[0, 0] = output[0] - (data[0, 0]/1000.0)**3
  # error[0, 0] = (output[0, 0]**3) / (1000**3) - (data[0, 0]**3) / (1000**3)
  # print(error)
  return error


# rebuild the female dataset using deform-based local method
def run(label="female"):
  start = time.time()
  pool = Pool(processes = 8)
  results = pool.map(rebuild, range(t_measure.shape[1]))
  pool.close()
  pool.join()

  wb = Workbook()
  ws = wb.get_active_sheet()
  for i in range(0, utils.M_NUM):
    ws.cell(row=1, column=i + 2).value = utils.M_STR[i]
  for i in range(len(results)):
    ws.cell(row=i + 2, column=1).value = i
    for j in range(0, results[i].shape[0]):
      ws.cell(row=i + 2, column=j + 2).value = results[i][j, 0]

  body_num = t_measure.shape[1]
  ans = np.array(results).reshape(body_num, utils.M_NUM).transpose()
  std = np.std(ans, axis=1)
  mean = np.mean(abs(ans), axis=1)
  ws.cell(row=body_num + 2, column=1).value = "mean error"
  ws.cell(row=body_num + 3, column=1).value = "std"
  for j in range(0, len(mean)):
    ws.cell(row=body_num + 2, column=j + 2).value = round(mean[j], 2)
    ws.cell(row=body_num + 3, column=j + 2).value = round(std[j],2)
  wb.save(os.path.join("../xlsx", "%s_%d_rebuild.xlsx"%(label, flag)))
  print("finish in %fs"%(time.time()-start))

if __name__ == "__main__":
  run(label)



        

