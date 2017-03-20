#!/usr/bin/python
# coding=utf-8

from .miner import *
from fancyimpute import *
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy.matlib
import numpy
import time
import math


# A Miner is responsible for mining the correlation between measure data human
# input:  all original data containing: measure...
# output: all correlation between measures
# usage:  impute the missing value in measure data
class Miner:

    def __init__(self, data):
        # load all necessary data
        self.test_size = 300
        self.data = data
        self.paras = self.data.paras
        self.flag_ = self.data.flag_

        self.data_path = self.paras['data_path'] + "miner/"
        self.ans_path = self.data.ans_path
        self.m2m = list(self.paras['m2m'])
        self.cfTable = self.itemCF()

        self.impute_name = ["SimpleFill", "SimpleAverage",
                            "CaseAmplifify", "KNN", "MICE"]
        self.impute_method = [SimpleFill(), self.simple_average,
                              self.case_amplify, KNN(), MICE()]

    # given flag, value, predict completed measures
    def get_predict(self, flag, data):
        if (flag == 1).sum() == self.data.m_num:
            return data
        else:
            solver = MICE()
            return self.imputate(flag, data, solver)

    # cal the similarity between measure, using Pearson Correlation(300-1531)
    def itemCF(self):
        print(' [**] begin get similarity of measures')
        start = time.time()
        cf_path = self.data_path + "cfTable_0%d.npy" % self.flag_
        cf_excel = self.ans_path + 'similarity_0%d.xlsx' % self.flag_
        table = numpy.zeros((self.data.m_num, self.data.m_num))
        if self.data.paras['reload_cfTable']:
            # calculation
            for i in range(0, self.data.m_num):
                table[i, i] = 0
                for j in range(i + 1, self.data.m_num):
                    a = numpy.array(self.data.t_measure[i, self.test_size:])
                    a.shape = (1, self.data.body_num - self.test_size)
                    b = numpy.array(self.data.t_measure[j, self.test_size:])
                    b.shape = (self.data.body_num - self.test_size, 1)
                    similar = numpy.linalg.norm(a) * numpy.linalg.norm(b)
                    similar = (a.dot(b) / similar)[0, 0]
                    table[i, j] = table[j, i] = similar
            # save the result in excel
            wb = Workbook()
            ws = wb.get_active_sheet()
            top = self.data.m_num
            for i in range(0, top):
                ws.cell(row=i * (top + 2) + 1, column=1).value = i
                ws.cell(row=i * (top + 2) + 1,
                        column=2).value = self.data.m_str[i]
                for j in range(0, top):
                    ws.cell(row=i * (top + 2) + j + 2,
                            column=1).value = self.data.m_str[j]
                    ws.cell(row=i * (top + 2) + j + 2,
                            column=2).value = table[i, j]
            wb.save(filename=cf_excel)
            numpy.save(open(cf_path, "wb"), table)
        else:
            table = numpy.load(open(cf_path, "rb"))
        print(' [**] finish load cfTable in %fs.' % (time.time() - start))
        return table

    # given flag value, predict completed measures, simple sum all
    def simple_average(self, flag, in_data):
        for j in range(0, in_data.shape[0]):
            if math.isnan(in_data[j, 0]):
                coeff = numpy.array(self.cfTable[j, :])
                coeff.shape = (self.data.m_num, 1)
                if (abs(coeff).sum() == 0):
                    in_data[j, 0] = 0
                else:
                    x = in_data[flag]
                    x.shape = (1, x.size)
                    coeff = coeff[flag]
                    in_data[j, 0] = x.dot(coeff) / (abs(coeff).sum())
        return in_data

    # Predict, let small smaller, big bigger
    def case_amplify(self, flag, in_data):
        for j in range(0, in_data.shape[0]):
            if math.isnan(in_data[j, 0]):
                coeff = numpy.array(self.cfTable[j, :])
                coeff.shape = (self.data.m_num, 1)
                if (abs(coeff).sum() == 0):
                    in_data[j, 0] = 0
                else:
                    x = in_data[flag]
                    x.shape = (1, x.size)
                    coeff = coeff[flag]
                    tmp = coeff * (abs(coeff)**1.5)
                    in_data[j, 0] = x.dot(tmp) / (abs(tmp).sum())
        return in_data

    # using imputation for missing data
    def imputate(self, flag, in_data, imputor):
        output = in_data.copy()
        output.shape = (self.data.m_num, 1)
        output[~flag] = numpy.nan
        if isinstance(imputor, type(self.simple_average)):
            output = imputor(flag, output)
        else:
            tmp = numpy.array(self.data.t_measure[:, self.test_size:])
            tmp = numpy.column_stack((tmp, output)).transpose()
            tmp = imputor.complete(tmp)
            output = numpy.array(tmp[-1, :]).reshape(self.data.m_num, 1)
        return output

    # test all methods
    def test(self):
        print(' [**] Begin predict input data...')
        table_size = 23
        input_file = self.data_path + "input.xlsx"
        input_wb = load_workbook(filename=input_file)
        sheets = input_wb.get_sheet_names()
        input_ws = input_wb.get_sheet_by_name(sheets[0])
        output_file = self.ans_path + '/prediction.xlsx'
        output_wb = Workbook()
        output_ws = output_wb.get_active_sheet()

        # read test data
        test_data = []
        for i in range(0, 5):
            flag = numpy.zeros((self.data.m_num, 1), dtype=bool)
            for j in range(0, self.data.m_num):
                flag[j, 0] = input_ws.cell(row=2 + j, column=2 + i).value
            test_data.append(flag)

        # begin test, using 0-300 data, output the offset to mean
        for i in range(0, len(test_data)):
            output_ws.cell(row=(i * table_size) + 2, column=1).value = i
            for j in range(0, self.data.m_num):
                output_ws.cell(row=(i * table_size) + 3 + j,
                               column=1).value = self.data.m_str[j]
            k = 2

            # test for imputation technology
            for m in range(0, len(self.impute_method)):
                output_ws.cell(row=(i * table_size) + 2,
                               column=k).value = self.impute_name[m]
                error = numpy.zeros((self.data.m_num, 1))
                for j in range(0, self.test_size):
                    print('Experiment %d, ' %
                          i, self.impute_name[m], ' sample: ', j)
                    in_data = numpy.array(self.data.t_measure[:, j])
                    in_data.shape = (self.data.m_num, 1)
                    out = self.imputate(
                        test_data[i], in_data, self.impute_method[m])
                    error += abs(out - in_data)
                error /= self.test_size
                error *= self.data.std_measure
                for j in range(0, self.data.m_num):
                    output_ws.cell(row=(i * table_size) + 3 + j,
                                   column=k).value = round(error[j, 0], 2)
                k = k + 1
                output_wb.save(output_file)
        output_wb.save(output_file)
        print(' [**] Finish predict input data...')


# test for this module
if __name__ == "__main__":
    male = MetaData("../parameter.json", 1)
    male_miner = Miner(male)
    male_miner.test()

    female = MetaData("../parameter.json", 2)
    female_miner = Miner(female)
    female_miner.test()
