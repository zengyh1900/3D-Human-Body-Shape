#!/usr/bin/python
# coding=utf-8

from openpyxl import Workbook
import scipy.sparse.linalg
import time
import numpy


# a model map measures to local face
class DeformLocal:

    def __init__(self, male, female):
        self.TYPE = "deform-local"
        self.body = [male, female]
        self.current_body = self.body[0]
        self.paras = self.current_body.paras

        self.demo_num = self.current_body.m_num
        self.ans_path = self.current_body.ans_path + "deform_local/"
        self.data_path = self.paras["data_path"]
        self.deformation = None
        self.L_list = self.local_matrix()

    def set_body(self, flag):
        self.current_body = self.body[flag - 1]

    # local map matrix: measure->deform
    def local_matrix(self):
        print(' [**] begin solve local_matrix')
        start = time.time()
        names = [self.data_path + "L_01.npy", self.data_path + "L_02.npy"]
        L_list = []
        if self.paras['reload_L']:
            for i, body in enumerate(self.body):
                L_tmp = []
                L_tosave = []
                for j in range(0, body.f_num):
                    print('  calc L_list for %d: NO.%d' % (body.flag_, j))
                    S = numpy.array(body.deform[:, j, :])
                    S.shape = (S.size, 1)
                    mask = numpy.array(body.mask[:, j])
                    mask.shape = (body.m_num, 1)
                    mask = mask.repeat(body.body_num, axis=1)
                    m = numpy.array(body.measure[mask])
                    m.shape = (m.size / body.body_num, body.body_num)
                    M = body.build_equation(m, 9)
                    # solve transform matrix
                    MtM = M.transpose().dot(M)
                    MtS = M.transpose().dot(S)
                    ans = numpy.array(scipy.sparse.linalg.spsolve(MtM, MtS))
                    ans.shape = (9, m.size / body.body_num)
                    # # -------------------------------------------
                    # for k in range(0, body.body_num):
                    #     s = numpy.array(body.deform[k, j, :])
                    #     s.shape = (9, 1)
                    #     m = numpy.array(body.measure[:, k][body.mask[:, k]])
                    #     m.shape = (m.size, 1)
                    #     res = ans.dot(m)
                    #     print("m:\n", m)
                    #     print("s:\n", s)
                    #     print("res:\n", res)
                    #     input()
                    # # -------------------------------------------
                    L_tmp.append(ans)
                    L_tosave.append(list(ans))
                L_list.append(L_tmp)
                numpy.save(open(names[i], "wb"), L_tosave)
        else:
            for fname in names:
                L_tmp = []
                tmp = numpy.load(open(fname, "rb"))
                for i in range(0, len(tmp)):
                    L_tmp.append(numpy.array([c for c in tmp[i]]))
                L_list.append(L_tmp)
        print(' [**] finish solve local_matrix in %fs' % (time.time() - start))
        return L_list

    # rebuild the female dataset using deform-based local method
    def local_rebuild(self):
        names = [self.ans_path + "01/", self.ans_path + "02/"]
        for i, body in enumerate(self.body):
            self.set_body(i + 1)
            error_path = self.ans_path + "d_local_0%d.xlsx" % (i + 1)
            error_npy = self.ans_path + "error_0%d.npy" % (i + 1)
            wb = Workbook()
            ws = wb.get_active_sheet()
            ans = numpy.zeros((body.m_num, body.body_num))
            for j in range(0, body.m_num):
                ws.cell(row=1, column=j + 2).value = body.m_str[j]
            for j in range(0, body.body_num):
                print('rebuilding deform_local-based: %d  ...' % j)
                ws.cell(row=j + 2, column=1).value = j
                data = body.t_measure[:, j].reshape(body.m_num, 1)
                [vertex, n, f] = self.mapping(data)
                body.save_obj(names[i] + body.file_list[j], vertex, f + 1)

                data = body.mean_measure + data * body.std_measure
                output = numpy.array(body.calc_measure(vertex))
                error = output - data
                error[0, 0] = (output[0, 0]**3) / (1000**3) - \
                    (data[0, 0]**3) / (1000**3)
                ans[:, j] = error.flat
                for k in range(0, error.shape[0]):
                    ws.cell(row=j + 2, column=k + 2).value = error[k, 0]
            std = numpy.std(ans, axis=1)
            mean = numpy.mean(abs(ans), axis=1)
            ws.cell(row=body.body_num + 2, column=1).value = "mean error"
            ws.cell(row=body.body_num + 3, column=1).value = "std"
            for j in range(0, len(mean)):
                ws.cell(row=body.body_num + 2, column=j + 2).value = mean[j]
                ws.cell(row=body.body_num + 3, column=j + 2).value = std[j]
            numpy.save(open(error_npy, "wb"), ans)
            wb.save(error_path)

    # given t_measures, return body shape
    def mapping(self, weight):
        weight = numpy.array(weight[:self.demo_num, :])
        weight.shape = (self.demo_num, 1)
        weight *= self.current_body.std_measure
        weight += self.current_body.mean_measure
        d = []
        L = self.L_list[self.current_body.flag_ - 1]
        for i in range(0, self.current_body.f_num):
            mask = numpy.array(self.current_body.mask[:, i])
            mask.shape = (self.current_body.m_num, 1)
            alpha = numpy.array(weight[mask])
            alpha.shape = (alpha.size, 1)
            s = L[i].dot(alpha)
            d += [a for a in s.flat]
        d = numpy.array(d)
        d.shape = (self.current_body.f_num * 9, 1)
        self.deformation = d
        return self.current_body.d_synthesize(d)
