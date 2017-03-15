#!/usr/bin/python
# coding=utf-8

from openpyxl import Workbook
import numpy
import time
import scipy


# a vertex-based global model, mainly for synthesis body from measure
# input: body measures
# usage: build a model in vertex-based pca space
# mapping measures to this space
# output: 3D human body shape
class VertexGlobal:

    def __init__(self, male, female):
        self.TYPE = "vertex-global"
        self.body = [male, female]
        self.current_body = self.body[0]
        self.paras = self.current_body.paras

        self.demo_num = self.current_body.m_num
        self.ans_path = self.current_body.ans_path + "vertex_global/"
        self.data_path = self.paras["data_path"]
        self.deformation = None
        self.m2v_ = self.load_m2v()

    def set_body(self, flag):
        self.current_body = self.body[flag - 1]

    # calculate the mapping matrix from measures to vertex-based
    def load_m2v(self):
        print(' [**] begin load_m2v ... ')
        start = time.time()
        m2v = []
        names = [self.data_path + "m2v_01.npy", self.data_path + "m2v_02.npy"]
        if self.paras["reload_m2v"]:
            for i, body in enumerate(self.body):
                V = body.v_coeff.transpose().copy()
                V.shape = (V.size, 1)
                M = body.build_equation(body.t_measure, body.v_basis_num)
                # solve transform matrix
                MtM = M.transpose().dot(M)
                MtV = M.transpose().dot(V)
                ans = numpy.array(scipy.sparse.linalg.spsolve(MtM, MtV))
                ans.shape = (body.v_basis_num, body.m_num)
                # ---------------------------------
                print("t_measure shape: \n", body.t_measure.shape)
                print("v_coeff shape: \n", body.v_coeff.shape)
                input()
                v = numpy.array(body.v_coeff[:, 0])
                v.shape = (body.d_basis_num, 1)
                m = numpy.array(body.t_measure[:, 0])
                m.shape = (body.m_num, 1)
                print(v.shape, m.shape, ans.shape)
                print("before:\n", v)
                print("after:\n", ans.dot(m))
                input()
                # ---------------------------------
                m2v.append(ans)
                numpy.save(open(names[i], "wb"), ans)
        else:
            for fname in names:
                tmp = numpy.load(open(fname), "rb")
                m2v.append(tmp)
        print(' [**] finish load_m2v  in %fs' % (time.time() - start))
        return m2v

    # rebuild the female dataset by vertex-global method
    def v_rebuild(self):
        names = [self.ans_path + "01/", self.ans_path + "02/"]
        for i, body in enumerate(self.body):
            self.set_body(i + 1)
            error_path = self.ans_path + "v_global_0%d.xlsx" % (i + 1)
            error_npy = self.ans_path + "error_0%d.npy" % (i + 1)
            wb = Workbook()
            ws = wb.get_active_sheet()
            ans = numpy.zeros((body.m_num, body.body_num))
            for j in range(0, body.m_num):
                ws.cell(row=1, column=j + 2).value = body.measure_str[j]
            for j in range(0, body.body_num):
                print('rebuilding vertex_global-based: %d  ...' % j)
                ws.cell(row=j + 2, column=1).value = j
                data = body.t_measure[:, j].reshape(body.m_num, 1)
                [vertex, n, f] = self.mapping(data)
                body.save_obj(names[i] + body.file_list[j], vertex, f + 1)

                data = body.mean_measure + body.std_measure * data
                output = numpy.array(body.calc_measure(vertex))
                error = output - data
                error[0, 0] = (output[0, 0]**3) / (1000**3) - \
                    (data[0, 0]**3) / (1000**3)
                ans[:, j] = error.flat
                for k in range(0, error.shape[0]):
                    ws.cell(row=j + 2, column=k + 2).value = error[k, 0]
                print(error)
            std = numpy.std(ans, axis=1)
            mean = numpy.mean(abs(ans), axis=1)
            ws.cell(row=body.body_num + 2, column=1).value = "mean error"
            ws.cell(row=body.body_num + 3, column=1).value = "std"
            for j in range(0, len(mean)):
                ws.cell(row=body.body_num + 2, column=j + 2).value = mean[j]
                ws.cell(row=body.body_num + 3, column=j + 2).value = std[j]
            numpy.save(open(error_npy, "wb"), ans)
            wb.save(error_path)

    # given measure, return body shape
    def mapping(self, weight):
        weight = numpy.array(weight[:self.demo_num, :])
        weight.shape = (self.demo_num, 1)

        m2v = self.m2v_[self.current_body.flag_ - 1]
        weight = m2v.dot(weight)
        [v, n, f] = self.current_body.v_synthesize(weight)
        # self.deformation = self.current_body.getDeform(v)
        return [v, n, f]
