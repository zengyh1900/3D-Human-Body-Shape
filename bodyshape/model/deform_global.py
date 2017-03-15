#!/usr/bin/python
# coding=utf-8

from openpyxl import Workbook
import scipy.sparse.linalg
import scipy.sparse
import scipy
import numpy
import time


# a class used to calculate corrlated massage about deform-based global method
# given: deform-based PCA coefficient, measure data
# output: the matrix trans measure into deform-based PCA space
class DeformGlobal:

    def __init__(self, male, female):
        self.TYPE = "deform-global"
        self.body = [male, female]
        self.current_body = self.body[0]
        self.paras = self.current_body.paras

        self.demo_num = self.current_body.m_num
        self.ans_path = self.current_body.ans_path + "deform_global/"
        self.data_path = self.paras["data_path"]
        self.deformation = None
        self.m2d_ = self.load_m2d()

    # calculate global mapping from measure to deformation PCA coeff
    def load_m2d(self):
        print(' [**] begin load_m2d ... ')
        start = time.time()
        m2d = []
        names = [self.data_path + "m2d_01.npy", self.data_path + "m2d_02.npy"]
        if self.data.paras['reload_m2d']:
            for i, body in enumerate(self.body):
                D = body.d_coeff.transpose().copy()
                D.shape = (D.size, 1)
                M = body.build_equation(body.t_measure, body.d_basis_num)
                # solve transform matrix
                MtM = M.transpose().dot(M)
                MtD = M.transpose().dot(D)
                ans = numpy.array(scipy.sparse.linalg.spsolve(MtM, MtD))
                ans.shape = (body.d_basis_num, body.m_num)
                m2d.append(ans)
                numpy.save(open(names[i], "wb"))
        else:
            for fname in names:
                tmp = numpy.load(open(fname, "rb"))
                m2d.append(tmp)
        print(' [**] finish load_m2d in %fs' % (time.time() - start))
        return m2d

    # rebuild the female dataset by deform-global method
    def d_rebuild(self):
        names = [self.ans_path + "01/", self.ans_path + "02/"]
        for i, body in enumerate(self.body):
            self.set_body(i + 1)
            error_path = self.ans_path + "d_global_0%d.xlsx" % (i + 1)
            error_npy = self.ans_path + "error_0%d.npy" % (i + 1)
            wb = Workbook()
            ws = wb.get_active_sheet()
            ans = numpy.zeros((body.m_num, body.body_num))
            for j in range(0, body.m_num):
                ws.cell(row=1, column=j + 2).value = body.measure_str[j]
            for j in range(0, body.body_num):
                print('rebuilding vertex_global-based: %d  ...' % j)
                ws.cell(row=j + 2, column=1).value = j
                data = body.t_measure[:, j].reshape(body.m_num, 1)
                [vertex, n, f] = self.mapping(data)
                body.save_obj(names[i] + body.file_list[j], vertex, f + 1)

                data = body.mean_measure + body.std_measure * data
                output = numpy.array(body.calc_measure(vertex))
                error = output - data
                error[0, 0] = (output[0, 0]**3) / (1000**3) - \
                    (data[0, 0]**3) / (1000**3)
                ans[:, j] = error.flat
                for k in range(0, error.shape[0]):
                    ws.cell(row=j + 2, column=k + 2).value = error[k, 0]
            std = numpy.std(ans, axis=1)
            mean = numpy.mean(abs(ans), axis=1)
            ws.cell(row=body.body_num + 2, column=1).value = "mean error"
            ws.cell(row=body.body_num + 3, column=1).value = "std"
            for j in range(0, len(mean)):
                ws.cell(row=body.body_num + 2, column=j + 2).value = mean[j]
                ws.cell(row=body.body_num + 3, column=j + 2).value = std[j]
            numpy.save(open(error_npy, "wb"), ans)
            wb.save(error_path)

    # given t_measure, return body shape
    def mapping(self, weight):
        weight = numpy.array(weight[:self.demo_num, :])
        weight.shape = (self.demo_num, 1)
        m2d = self.m2d_[self.current_body.flag_ - 1]
        weight = m2d.dot(weight)

        basis = self.current_body.d_basis[:, :self.current_body.d_basis_num]
        d = numpy.matmul(basis, weight)
        [v, n, f] = self.current_body.d_synthesize(d)
        return [v, n, f]